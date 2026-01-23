import io
from typing import List, Dict, Any

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.enums import TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def markdown_to_pdf_bytes(md_text: str) -> bytes:
    """Converte um texto Markdown em PDF básico (títulos, listas e blocos de código).
    Retorna bytes do PDF para envio como resposta HTTP.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    h3 = styles["Heading3"]

    code_style = ParagraphStyle(
        name="Code",
        parent=body,
        fontName="Courier",
        fontSize=9,
        leading=11,
        alignment=TA_LEFT,
    )

    story = []
    in_code_block = False
    code_lines: List[str] = []

    def flush_code_block():
        nonlocal code_lines
        if code_lines:
            # Escapa sinais de < > para não quebrar markup do Paragraph
            safe_lines = [l.replace("<", "&lt;").replace(">", "&gt;") for l in code_lines]
            txt = "<font face='Courier'>" + "<br/>".join(safe_lines) + "</font>"
            story.append(Paragraph(txt, code_style))
            story.append(Spacer(1, 6))
            code_lines = []

    for raw_line in md_text.splitlines():
        line = raw_line.rstrip()
        # Detecta início/fim de bloco de código
        if line.strip().startswith("```"):
            if in_code_block:
                # Finaliza bloco de código
                flush_code_block()
                in_code_block = False
            else:
                # Inicia bloco de código
                in_code_block = True
            continue

        if in_code_block:
            code_lines.append(line)
            continue

        # Títulos
        if line.startswith("# "):
            story.append(Paragraph(line[2:].strip(), h1))
            story.append(Spacer(1, 8))
            continue
        if line.startswith("## "):
            story.append(Paragraph(line[3:].strip(), h2))
            story.append(Spacer(1, 6))
            continue
        if line.startswith("### "):
            story.append(Paragraph(line[4:].strip(), h3))
            story.append(Spacer(1, 6))
            continue

        # Listas simples
        if line.startswith("- "):
            story.append(Paragraph("• " + line[2:].strip(), body))
            story.append(Spacer(1, 4))
            continue

        # Parágrafos comuns (inclui linhas em branco)
        if line.strip() == "":
            story.append(Spacer(1, 6))
        else:
            story.append(Paragraph(line, body))
            story.append(Spacer(1, 6))

    # Se terminar o arquivo dentro de um bloco de código
    if in_code_block:
        flush_code_block()

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def gerar_excel_relatorio(md_text: str, dados_relatorio: List[Dict[str, Any]], nome_planilha_dados: str = "Dados") -> bytes:
    """Gera um arquivo Excel com:
    - Aba "Relatorio" contendo o Markdown gerado pela IA (uma linha por parágrafo)
    - Aba com os dados tabulares do relatório (se disponíveis)
    Retorna bytes do arquivo XLSX para resposta HTTP.
    """
    wb = Workbook()

    # Aba do relatório em Markdown
    ws_rel = wb.active
    ws_rel.title = "Relatorio"
    ws_rel.column_dimensions["A"].width = 120
    title_font = Font(size=12, bold=False)
    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, line in enumerate(md_text.splitlines(), start=1):
        cell = ws_rel.cell(row=idx, column=1, value=line)
        cell.font = title_font
        cell.alignment = wrap

    # Aba de dados
    ws_dados = wb.create_sheet(title=nome_planilha_dados)
    if dados_relatorio:
        headers = list(dados_relatorio[0].keys())
        ws_dados.append(headers)
        for row in dados_relatorio:
            ws_dados.append([row.get(h) for h in headers])

        # Ajusta largura básica das colunas
        for i, h in enumerate(headers, start=1):
            ws_dados.column_dimensions[chr(64 + i)].width = max(12, min(40, len(str(h)) + 4))

    # Salva em bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()